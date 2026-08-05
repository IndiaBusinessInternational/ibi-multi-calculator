#!/usr/bin/env python3
"""Rebuild catalogue.json from the IBI product master held in Google Drive.

The master is link-shared, so no credentials are needed -- the plain
`uc?export=download` endpoint returns the workbook bytes. Run by
.github/workflows/sync-catalogue.yml; it only rewrites catalogue.json, and the
workflow only commits when the content actually changed.

Every sanity check below is a hard failure on purpose: if Drive serves an error
page, a sign-in page or a truncated file, this must exit non-zero rather than
publish a short or empty catalogue that would wipe the picker in the app.
"""

import io
import json
import os
import sys
import urllib.request
import zipfile
from datetime import datetime, timezone

FILE_ID = '18_gEDm7vNZt1-skIZqxu-FyDPxGUO9Yw'
SOURCE_NAME = 'Product Names, HSN Code, GST, Prod & Pack Dimensions 5 Aug 2026.xlsx'

# The catalogue is the FIRST tab only. The workbook carries a second tab of
# working notes that must never reach the apps. Sheet names are capped at 31
# characters by the file format and the user re-dates this one on every
# revision ("... 5 Aug 2026"), so match a stable prefix, never the full title.
SHEET_PREFIX = 'Product Names'

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'catalogue.json')

MIN_PRODUCTS = 200      # the master has 268; a big drop means something went wrong
MAX_PRODUCTS = 600


def download(file_id):
    url = 'https://drive.google.com/uc?export=download&id=' + file_id
    req = urllib.request.Request(url, headers={'User-Agent': 'ibi-catalogue-sync'})
    with urllib.request.urlopen(req, timeout=90) as r:
        data = r.read()
    if not data.startswith(b'PK'):
        sys.exit('ERROR: Drive did not return an xlsx (first bytes: %r). '
                 'Is the file still shared with "anyone with the link"?' % data[:40])
    if not zipfile.is_zipfile(io.BytesIO(data)):
        sys.exit('ERROR: downloaded bytes are not a readable zip/xlsx.')
    return data


def norm_hsn(v):
    """HSN codes are 8 digits. Excel stores some as numbers, which silently eats
    a leading zero (07099990 -> 7099990); 33 of the 268 codes start with one."""
    if v is None:
        return ''
    if isinstance(v, float) and v == int(v):
        v = int(v)
    s = str(v).strip()
    if s.isdigit() and len(s) % 2 == 1:
        s = '0' + s
    return s


def norm_gst(v):
    """GST arrives two ways in this workbook: as a real percentage-formatted
    number (0.05 means 5%) and as free text ("0% (NIL)"). Scraping the digits
    out of the text form happens to work for whole percentages but turns 0.025
    into 25, so branch on the type instead."""
    if v is None:
        return 0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        pct = float(v) * 100 if abs(float(v)) <= 1 else float(v)
    else:
        digits = ''
        for ch in str(v).split('%')[0]:
            if ch.isdigit() or (ch == '.' and '.' not in digits):
                digits += ch
        if not digits:
            return 0
        pct = float(digits)
    pct = round(pct, 2)
    return int(pct) if pct == int(pct) else pct


def num(v):
    """A measurement cell -> float, or None when blank. About a fifth of the
    rows have no dimensions recorded yet, and a missing measurement must stay
    missing rather than become a zero the apps would treat as real."""
    if v is None or str(v).strip() == '':
        return None
    try:
        f = float(str(v).strip())
    except ValueError:
        return None
    return int(f) if f == int(f) else f


def find_columns(rows):
    """Map header text -> column index. The layout gained six columns between
    the July and August masters, so read the positions off the header row
    rather than hard-coding them."""
    want = {
        'name':  ('product name',),
        'hsn':   ('hsn code', 'hsn'),
        'gst':   ('gst rate', 'gst'),
        'cat':   ('category',),
        'pl':    ('prod l',), 'pw': ('prod w',), 'ph': ('prod h',), 'pwt': ('prod wt',),
        'kl':    ('pkg l',),  'kw': ('pkg w',),  'kh': ('pkg h',),  'kwt': ('pkg wt',),
    }
    for row in rows[:20]:
        cells = {}
        for i, c in enumerate(row or ()):
            if c is None:
                continue
            cells[str(c).strip().lower()] = i
        if 'product name' not in cells:
            continue
        got = {}
        for key, aliases in want.items():
            for a in aliases:
                hit = next((i for text, i in cells.items() if text.startswith(a)), None)
                if hit is not None:
                    got[key] = hit
                    break
        if 'name' in got and 'hsn' in got:
            return got
    sys.exit('ERROR: could not find a header row with "Product Name" + "HSN".')


def parse(data):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)

    sheet = wb.sheetnames[0]
    if not sheet.lower().startswith(SHEET_PREFIX.lower()):
        sys.exit('ERROR: first sheet is %r, expected one starting %r. Sheets: %s'
                 % (sheet, SHEET_PREFIX, wb.sheetnames))
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    col = find_columns(rows)

    def cell(row, key):
        i = col.get(key)
        return row[i] if i is not None and len(row) > i else None

    out = []
    for row in rows:
        if not row or row[0] is None:
            continue
        try:                                    # data rows carry a numeric S.No;
            int(float(str(row[0]).strip()))     # title lines and the "> CATEGORY"
        except (TypeError, ValueError):         # banner rows do not
            continue
        name = str(cell(row, 'name')).strip() if cell(row, 'name') else ''
        if not name:
            continue
        dims = [num(cell(row, k)) for k in ('pl', 'pw', 'ph')]
        pkg = [num(cell(row, k)) for k in ('kl', 'kw', 'kh')]
        out.append({
            'n': name,
            'h': norm_hsn(cell(row, 'hsn')),
            'g': norm_gst(cell(row, 'gst')),
            'c': str(cell(row, 'cat')).strip() if cell(row, 'cat') else '',
            'd': dims if all(v is not None for v in dims) else None,
            'wt': num(cell(row, 'pwt')),
            'pd': pkg if all(v is not None for v in pkg) else None,
            'pwt': num(cell(row, 'kwt')),
        })
    return out


def main():
    products = parse(download(FILE_ID))

    if not MIN_PRODUCTS <= len(products) <= MAX_PRODUCTS:
        sys.exit('ERROR: parsed %d products, expected %d-%d. Refusing to publish.'
                 % (len(products), MIN_PRODUCTS, MAX_PRODUCTS))
    names = [p['n'] for p in products]
    dupes = sorted({n for n in names if names.count(n) > 1})
    if dupes:
        sys.exit('ERROR: duplicate product names in the master: %s' % dupes[:5])
    odd = [p for p in products if p['h'] and len(p['h']) != 8]
    if odd:
        print('WARNING: %d HSN codes are not 8 digits, e.g. %s'
              % (len(odd), [(p['n'][:40], p['h']) for p in odd[:3]]))
    nodims = sum(1 for p in products if not p['d'])
    if nodims:
        print('NOTE: %d of %d products have no product dimensions recorded'
              % (nodims, len(products)))

    products.sort(key=lambda p: p['n'].lower())
    payload = {
        'generated': datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
        'source': SOURCE_NAME,
        'fileId': FILE_ID,
        'count': len(products),
        'products': products,
    }

    old = None
    if os.path.exists(OUT):
        try:
            with io.open(OUT, encoding='utf-8') as f:
                old = json.load(f)
        except (ValueError, OSError):
            old = None

    # `generated` moves on every run, so rewriting unconditionally would produce
    # a commit every five minutes. Only touch the file when a product changed.
    changed = not old or old.get('products') != products
    if not changed:
        print('%d products, no change — catalogue.json left untouched' % len(products))
        return

    tmp = OUT + '.tmp'
    with io.open(tmp, 'w', encoding='utf-8', newline='\n') as f:
        f.write(json.dumps(payload, ensure_ascii=False, indent=1) + '\n')
    os.replace(tmp, OUT)

    print('%d products -> catalogue.json UPDATED' % len(products))
    if old:
        was, now = {p['n'] for p in old.get('products', [])}, set(names)
        for n in sorted(now - was)[:40]:
            print('   + %s' % n)
        for n in sorted(was - now)[:40]:
            print('   - %s' % n)


if __name__ == '__main__':
    main()
